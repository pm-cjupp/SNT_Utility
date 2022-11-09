"""--------------------------------------------------------------------------------------------------------------------
SNT_Module: Final version of the functions to be used for the Serial Number Tool GUI. This tool is intended for use
in Planar Motor Inc. production operations

Created by:     Cameron Jupp
Date Started:   November 2, 2022
--------------------------------------------------------------------------------------------------------------------"""
# -------------------------------------------------------------------------------------------------------------------- #
# -------------------------------------  /^\  / /  / __  /  /__ __/  / ___/ ------------------------------------------ #
# ------------------------------------  / /\\/ /  / /_/ /    / /    / __/  ------------------------------------------- #
# -----------------------------------  /_/  \_/  /_____/    /_/    /____/ -------------------------------------------- #
# -------------------------------------------------------------------------------------------------------------------- #
# Last left off at:
#
# - Add names for columns and sheets
#
# - Code auto indexing function
#
# - Add "download sheet" and "update changes" buttons to GUI
#
# - Create new format for GUI on tablet
#
# - Connect any worksheet functions with sheets object
#
#
# -------------------------------------------------------------------------------------------------------------------- #

# -------------------------- Imported Modules -----------------------------------

import openpyxl
from synology_drive_api.drive import SynologyDrive


# -------------------------- Global Variables -----------------------------------

sheets_url = "https://docs.google.com/spreadsheets/d/1Ikp-DhcNI44rnxT0W_aPSAQksGQ8fAG9VZoKgV0_bkU/edit?usp=sharing"


header_row = 2

br_fw_sn_prefix    ='22318'
pmi_fw_sn_prefix ='22310'

controller_sn_prefix = '95231'

body_id_col = 0

fw_sn_col = 7
fw_amp_sn_col = 24
fw_cont_sn_col = 25

NAS_IP = '192.168.1.65'
NAS_PORT = '5001'

#NAS_USER = 'SharedAccount'
#NAS_PASS = 'pQS3s6BPPn7u'

NAS_USER = 'cjupp'
NAS_PASS = 'ZJcj4GBSF9Frr9KCmGxy'

prod_wb_name = 'Production_Record_Testbench.osheet'
file_path = '/team-folders/cjupp/Projects/Serial Number Tool/'

prod_ss = 'ProductionScreenshot.xlsx'


# ------------------------------- Classes ----------------------------------------
class Sheet:
    def __init__(self, sheet_obj, header_list, header_row):
        # -------------------------------------------------------------------
        #Define the number of rows
        self.row_cnt = sheet_obj.max_row
        #Define the number of columns
        self.col_cnt = sheet_obj.max_column

        #Store the header data
        self.headers = header_list
        self.header_row = header_row

        #Transfer the data into a list of lists for simple use
        self.get_data(sheet_obj)

        #Automatically find the header names and save them
        self.get_headers()
        #-------------------------------------------------------------------

    def get_data(self, sheet_obj):
        # -------------------------------------------------------------------
        self.data = []
        for row in range(1, self.row_cnt):
            row_data = []
            for col in range(1, self.col_cnt):
                # print(sheet_obj.cell(row, col).value)
                row_data.append(sheet_obj.cell(row, col).value)
            # print(row_data)
            self.data.append(row_data)
         # -------------------------------------------------------------------

    def get_headers(self):
        # -------------------------------------------------------------------
        self.header_indexes = []
        #For every column in the header row
        for cols in range(self.col_cnt):

            #For every possible name of a given header
            for header_lists in self.headers:

                #If one of the possible names matches
                if header_lists == self.data[self.header_row][cols]:
                    #Save the name in a list as an index for use by other parts of the program
                    self.header_indexes.append(header_lists)
        #-------------------------------------------------------------------

class InfoLine:
    def __init__(self, info, line, default):
        self.line = self.FindLineNum()
        self.info = info
        self.line = line
        self.default = default

    def FindLineNum(self):
        pass

    def AppendInfo(self, newInfo):
        self.info.append(newInfo)

    def RestoreInfo(self):
        self.info = self.default


class Flyway:
    def __init__(self, sn, controller, amp):
        self.sn         = sn
        self.controller = controller
        self.amp        = amp

# -------------------------------------------------------------------------------------------------------------------- #
# ----------------------  / ___/  / / //  /^\  / /  / ___/ /__  __/  /_  _/  / __  /  /^\  / / ----------------------- #
# ---------------------  / __/   / /_//  / /\\/ /  / /__     / /      / /   / /_/ /  / /\\/ / ------------------------ #
# --------------------  /_/     /____/  /_/  \_/  /____/    /_/    /____/  /_____/  /_/  \_/ ------------------------- #
# -------------------------------------------------------------------------------------------------------------------- #

"""------------------------------------------------------------------------------------
init_sheets: calls all initialization functions needed for the program to run properly
                -----------------------------------------------
Arguments:
 - 
                -----------------------------------------------
Returns: nothing
                -----------------------------------------------
Created by: Cameron Jupp
Date:       Aug 24, 2022
Edited:
------------------------------------------------------------------------------------"""
def init_sheets():
    #Download the worksheet from Synology
    #download_sheets()

    #Load the worksheet from local path
    prod_wb = openpyxl.load_workbook(prod_ss)

    #Create a dictionary for the sheets
    prod_sheets = {}

    #Assign dictionary entries based on sheet names, then create sheet objects from them
    for sheet in range(len(prod_wb.sheetnames)):
        prod_sheets[prod_wb.sheetnames[sheet]] = Sheet(prod_wb.worksheets[sheet], [], 2)

    #Troubleshooting statements
    #print(prod_sheets)
    #print(prod_sheets["ERP"].col_cnt)

    #Return the sheets dictionary to main for use
    return prod_sheets


"""------------------------------------------------------------------------------------
download_sheets: Downloads the sheets file from synology drive
                -----------------------------------------------
Arguments:
 - 
                -----------------------------------------------
Returns: nothing
                -----------------------------------------------
Created by: Cameron Jupp
Date:       , 2022
Edited:
------------------------------------------------------------------------------------"""
def download_sheets():
    with SynologyDrive(NAS_USER, NAS_PASS, NAS_IP, NAS_PORT) as synd:

        file = synd.download_synology_office_file(file_path + prod_wb_name)
        with open(prod_ss, 'wb') as f:
            f.write(file.read())
    pass


"""------------------------------------------------------------------------------------
update_sheets: Uploads an edited sheets file back into synology drive
                -----------------------------------------------
Arguments:
 -
                -----------------------------------------------
Returns: nothing
                -----------------------------------------------
Created by: Cameron Jupp
Date:       , 2022
Edited:
------------------------------------------------------------------------------------"""
def update_sheets():
    with SynologyDrive(NAS_USER, NAS_PASS, NAS_IP, NAS_PORT) as synd:
        with open(prod_ss, 'rb') as file:
            synd.upload_file(file, file_path)
    pass


"""------------------------------------------------------------------------------------
init_headers: initializes the headers of the worksheets with their names and columns
                -----------------------------------------------
Arguments:
 - 
                -----------------------------------------------
Returns: nothing
                -----------------------------------------------
Created by: Cameron Jupp
Date:       Aug 24, 2022
Edited:
------------------------------------------------------------------------------------"""
def init_headers(worksheet):
    headers = worksheet[header_row]
    return headers

"""------------------------------------------------------------------------------------
search_wks: searches an entire worksheet for a particular string
                -----------------------------------------------
Arguments:
 - search_string: the string to be searched for
 
 - worksheet: the worksheet object to be searched within
                -----------------------------------------------
Returns: 
- "none_found" if no matching string is found

- "multiple_found" if more than one result is found 

- otherwise, if one instance is found, it returns the number of the row it was found at as an integer
        
                -----------------------------------------------
Created by: Cameron Jupp
Date:       Aug 24, 2022
Edited:
------------------------------------------------------------------------------------"""
def search_wks(search_string, worksheet):
    #search entire column
    #if a match is found, log the row it was found in and continue searching, but set a flag that the value was found
    #if another match is found, trigger another flag to indicate a repeat value
    #return the row that

    cell_found = 0
    result = []
    col_indx = 1
    row_indx = 1

    for row in worksheet.data:
        for item in row:
            value = item
            if value == search_string:
                cell_found += 1
                result.append(row_indx)
                result.append(col_indx)
            col_indx+=1
        row_indx+=1
        col_indx=1

    if cell_found == 0:
        return "none_found"
    elif cell_found == 1:
        return result
    else:
        return "multiple_found"

"""------------------------------------------------------------------------------------
search_col: 
                -----------------------------------------------
Arguments:
 - search_string: the string to be searched
 
 - column: the column of choice to search in
 
 - worksheet: the worksheet object the string is being search for in
                -----------------------------------------------
Returns: 
- "none_found" if no matching string is found

- "multiple_found" if more than one result is found 

- otherwise, if one instance is found, it returns the number of the row it was found at as an integer
                -----------------------------------------------
Created by: Cameron Jupp
Date:       Aug 24, 2022
Edited:
------------------------------------------------------------------------------------"""
def search_col(search_string, column, worksheet):
    #search entire column
    #if a match is found, log the row it was found in and continue searching, but set a flag that the value was found
    #if another match is found, trigger another flag to indicate a repeat value
    #return the row that

    cell_found = 0
    row_result = 0
    row_indx = 1

    for row in worksheet.data:
        value = row[column]
        #print(value)
        if value == search_string:
            cell_found += 1
            row_result = row_indx
            #print(cell_found)
        row_indx+=1

    if cell_found == 0:
        return "none_found"
    elif cell_found == 1:
        return row_result
    else:
        return "multiple_found"

"""------------------------------------------------------------------------------------
get_highest_sn: 
                -----------------------------------------------
Arguments:
 - worksheet: the worksheet object that the sn is being searched within
 
 - wks_size: the size of the worksheet, to prevent unecessary extra calculations
 
 - column: the column that the serial number type of choice is in
 
 - company: the company the flyway is being built for. This impacts the layout of the 
 serial number
 
 - component: the type of part the serial number is for
                -----------------------------------------------
Returns: nothing
                -----------------------------------------------
Created by: Cameron Jupp
Date:       Aug 24, 2022
Edited:
------------------------------------------------------------------------------------"""
def get_highest_sn(worksheet, column, company, component):
    #print(wks_size)
    #print("style" + company)
    highest_sn = 0
    sn_prefix = '0'
    if component == 'controller':
       sn_prefix = controller_sn_prefix

    elif component == 'flyway':
        if company == "1":
            sn_prefix = br_fw_sn_prefix
        else:
            sn_prefix = pmi_fw_sn_prefix

    elif component == 'body':
        if company == 'B&R':
            sn_prefix = ''
        else:
            sn_prefix = ''


    for row in range(2, worksheet.row_cnt-1):
        print(type(row))
        print(type(column))
        print(column)
        print("row:" + str(row))
        print(worksheet.row_cnt)
        print(worksheet.col_cnt)
        print( "worksheet cell:" + str(worksheet.data[row][column]))
        cell_string = str(worksheet.data[row][column])

        if cell_string != '':
            #print("cell string:" + cell_string[0:5])

            if cell_string[0:5] == sn_prefix:
                cell_value = int(cell_string, 16)
                #print("cell value:" + str(cell_value))

                if int(cell_value) > int(highest_sn):       #FIX VAR TYPE ISSUE HERE
                    highest_sn = int(cell_string, 16)
                    #print("highest sn:" + str(highest_sn))
    print(highest_sn)
    print(hex(highest_sn))
    return highest_sn


"""------------------------------------------------------------------------------------
add_flyway: adds a new flyway serial number to the worksheet based on the observed highest, 
then also adds in the amp and controller serial numbers input by the user
                -----------------------------------------------
Arguments:
 - worksheet: the worksheet object that the information being added lives on
 
 - body_id: the body id of the flyway being built
 
 - amp_sn: the serial number of the amp being added to the flyway
 
 - controller_sn: the serial number of the controller being added to the flyway
 
 - company: the company the flyway is being built for, affects the type of serial number
 that will be searched for and added
                -----------------------------------------------
Returns: nothing
                -----------------------------------------------
Created by: Cameron Jupp
Date:       Aug 24, 2022
Edited:
------------------------------------------------------------------------------------"""
def add_flyway(worksheet, body_id, amp_sn, controller_sn, company):
    highest_sn = 0
    next_free_row = 0

    #Check to see if the body ID input already exists
    if search_col(body_id, 1, worksheet.data) == "none_found":
        wks_len = worksheet.row_cnt
        print(wks_len)

    else:
        return "existing_id"

    #Find the value of the highest serial number for a given company
    highest_sn = get_highest_sn(worksheet, fw_sn_col, company, "flyway")
    print(highest_sn)

    #Get the row of the Body ID you want to write to
    fw_row = search_col(body_id, body_id_col, worksheet)
    print(fw_row)

    #Increment and write the value to the next row
    print(hex(highest_sn + 1).upper())
    new_fw_sn = hex(highest_sn + 1).upper()
    new_fw_sn = new_fw_sn[2:]
    worksheet.update_value((fw_row, fw_sn_col+1), new_fw_sn)
    print("SN updated")

    #Add the amp and controller serial numbers
    worksheet.data[fw_row][fw_amp_sn_col] = amp_sn
    print("Amp SN updated")
    worksheet.data[fw_row][fw_amp_sn_col] = controller_sn
    print("Controller SN updated")

"""------------------------------------------------------------------------------------
:
                -----------------------------------------------
Arguments:
 - 
                -----------------------------------------------
Returns: nothing
                -----------------------------------------------
Created by: Cameron Jupp
Date:       , 2022
Edited:
------------------------------------------------------------------------------------"""
def add_controller(worksheet, controller_sn):
    return 0


#--------------------------------------------------------------------------------------------------------------------- #
# -------------------------------------  /^\/^\      /^^\    |_ _|    /^\  / /  -------------------------------------- #
# ------------------------------------  / /\/\ \    / /_\\    | |    / /\\/ /  --------------------------------------- #
# -----------------------------------  /_/    \_\  /_/   \\  |___|  /_/  \_/  ---------------------------------------- #
#----------------------------------------------------------------------------------------------------------------------#

if __name__ == "__main__":
    #Initialize the sheet, worksheets, and any other prerequisites
    sheets_array = init_sheets()
    for rows in sheets_array["Production Log"].data:
        #print(rows)
        for cells in rows:
            print(cells)











